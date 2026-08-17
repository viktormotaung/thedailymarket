from decimal import Decimal, ROUND_HALF_UP
import re

from openpyxl import load_workbook
from django.db import transaction
from django.core.exceptions import ValidationError

from products.models import (
    Category,
    Product,
    ProductPricing,
    ProductPriceHistory,
)
from suppliers.models import Supplier


# ======================================================
# HELPERS
# ======================================================

def _vat_included(val):
    return str(val).strip().lower() in ("yes", "true", "1")


def _is_active(val):
    return str(val).strip().lower() in ("yes", "true", "1")


def _dec(val):
    try:
        if val in (None, ""):
            return Decimal("0.00")

        return Decimal(str(val)).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP
        )

    except Exception:
        return Decimal("0.00")


def generate_supplier_code(name: str, db="default") -> str:
    base = re.sub(r"[^A-Z0-9]", "", name.upper())[:10] or "SUPPLIER"
    code = base
    counter = 1

    while Supplier.objects.using(db).filter(code=code).exists():
        counter += 1
        code = f"{base[:7]}{counter}"

    return code


# ======================================================
# MAIN IMPORTER
# ======================================================

def import_products_from_excel(file, db="default"):
    """
    DB-AWARE IMPORTER
    """

    # ======================================================
    # STEP 1 — LOAD FILE
    # ======================================================
    try:
        wb = load_workbook(file, data_only=True)
    except Exception as e:
        raise ValidationError(f"STEP 1 (File load): {e}")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 3:
        raise ValidationError("STEP 1: Excel must contain at least 3 rows")

    # ======================================================
    # STEP 2 — HEADERS
    # ======================================================
    HEADER_ROW = 1
    DATA_START = 2

    headers = rows[HEADER_ROW]
    if not headers:
        raise ValidationError("STEP 2: Header row is empty")

    COL = {
        "category": 0,
        "subcategory": 1,
        "product": 2,
        "product_no": 3,
        "description": 4,
        "supplier": 5,
        "uom": 6,
        "vat_included": 8,
        "cost_price": 9,
        "wholesale_margin": 12,
        "retail_margin": 15,
        "is_active": 18,
    }

    EXPECTED_HEADERS = {
        0: "Category",
        1: "Subcategory",
        2: "Product",
        3: "Product No.",
        5: "Supplier",
        6: "UOM",
        8: "Vat Included",
        9: "Price",
        12: "Profit (%)",
        15: "Profit (%)",
        18: "Is Active",
    }

    for idx, expected in EXPECTED_HEADERS.items():
        actual = str(headers[idx]).strip() if headers[idx] else ""
        if actual != expected:
            raise ValidationError(
                f"STEP 2: Expected '{expected}' in column {idx + 1}, found '{actual}'"
            )

    created = 0
    updated = 0
    unchanged = 0
    price_changes = 0

    # ======================================================
    # STEP 3 — PROCESS ROWS
    # ======================================================
    for excel_row, row in enumerate(rows[DATA_START:], start=DATA_START + 1):

        if not row or not row[COL["product_no"]]:
            continue

        try:
            with transaction.atomic(using=db):

                # -----------------------------
                # READ DATA
                # -----------------------------
                category_name = str(row[COL["category"]]).strip()
                subcategory_name = str(row[COL["subcategory"]]).strip()
                product_name = str(row[COL["product"]]).strip()
                product_no = str(row[COL["product_no"]]).strip()
                description = row[COL["description"]] or ""

                supplier_name = str(row[COL["supplier"]]).strip()
                if not supplier_name:
                    raise ValidationError("Supplier name is required")

                raw_uom = row[COL["uom"]]
                uom = str(raw_uom).strip().upper() if raw_uom else "EA"
                if uom not in {"EA", "KG", "L", "PK", "BOX"}:
                    raise ValidationError(f"Invalid UOM '{uom}'")

                vat_included = _vat_included(row[COL["vat_included"]])

                cost_price = _dec(row[COL["cost_price"]])

                wholesale_margin = str(int(float(row[COL["wholesale_margin"]] or 0)))
                retail_margin = str(int(float(row[COL["retail_margin"]] or 0)))

                is_active = _is_active(row[COL["is_active"]])

                # -----------------------------
                # CATEGORY
                # -----------------------------
                parent_cat, _ = Category.objects.using(db).get_or_create(
                    name=category_name,
                    parent=None,
                )

                sub_cat, _ = Category.objects.using(db).get_or_create(
                    name=subcategory_name,
                    parent=parent_cat,
                )

                # -----------------------------
                # PRODUCT
                # -----------------------------
                product, was_created = Product.objects.using(db).update_or_create(
                    product_no=product_no,
                    defaults={
                        "name": product_name,
                        "category": sub_cat,
                        "description": description,
                        "uom": uom,
                        "cost_price": cost_price,
                    },
                )

                created += int(was_created)
                updated += int(not was_created)

                # -----------------------------
                # SUPPLIER
                # -----------------------------
                supplier = Supplier.objects.using(db).filter(name=supplier_name).first()

                if not supplier:
                    supplier = Supplier.objects.using(db).create(
                        name=supplier_name,
                        code=generate_supplier_code(supplier_name, db),
                        is_active=True,
                    )

                supplier.categories.add(parent_cat)

                # -----------------------------
                # PRICING
                # -----------------------------

                pricing = (
                    ProductPricing.objects.using(db)
                    .filter(
                        product=product,
                        supplier=supplier,
                    )
                    .first()
                )

                pricing_was_created = pricing is None

                # Capture OLD values before changing anything.
                if pricing_was_created:
                    old_cost = None
                    old_wholesale = None
                    old_retail = None
                else:
                    old_cost = pricing.supplier_price_incl
                    old_wholesale = pricing.wholesale_price_excl
                    old_retail = pricing.retail_price_excl


                if pricing_was_created:
                    pricing = ProductPricing(
                        product=product,
                        supplier=supplier,
                    )

                pricing.supplier_price_input = cost_price
                pricing.supplier_price_is_inclusive = vat_included
                pricing.wholesale_margin_percent = wholesale_margin
                pricing.retail_margin_percent = retail_margin
                pricing.is_active = is_active

                # Save so ProductPricing calculates its canonical supplier price
                # and its calculated wholesale / retail prices.
                pricing.save(using=db)

                # NEW values after save.
                new_cost = pricing.supplier_price_incl
                new_wholesale = pricing.wholesale_price_excl
                new_retail = pricing.retail_price_excl


                # -----------------------------
                # PRICE CHANGE DETECTION
                # -----------------------------

                if pricing_was_created:
                    ProductPriceHistory.objects.using(db).create(
                        product=product,
                        supplier=supplier,
                        change_type="INITIAL",

                        previous_cost_price=None,
                        new_cost_price=new_cost,
                        cost_change=None,

                        previous_wholesale_price=None,
                        new_wholesale_price=new_wholesale,
                        wholesale_change=None,

                        previous_retail_price=None,
                        new_retail_price=new_retail,
                        retail_change=None,
                    )

                else:
                    cost_changed = old_cost != new_cost
                    wholesale_changed = old_wholesale != new_wholesale
                    retail_changed = old_retail != new_retail

                    if cost_changed or wholesale_changed or retail_changed:

                        cost_change = (
                            new_cost - old_cost
                            if cost_changed
                            else Decimal("0.00")
                        )

                        wholesale_change = (
                            new_wholesale - old_wholesale
                            if wholesale_changed
                            else Decimal("0.00")
                        )

                        retail_change = (
                            new_retail - old_retail
                            if retail_changed
                            else Decimal("0.00")
                        )

                        # Determine overall direction.
                        changes = [
                            cost_change,
                            wholesale_change,
                            retail_change,
                        ]

                        positive = any(change > 0 for change in changes)
                        negative = any(change < 0 for change in changes)

                        if positive and negative:
                            change_type = "MIXED"
                        elif positive:
                            change_type = "INCREASE"
                        else:
                            change_type = "DECREASE"

                        ProductPriceHistory.objects.using(db).create(
                            product=product,
                            supplier=supplier,
                            change_type=change_type,

                            previous_cost_price=old_cost,
                            new_cost_price=new_cost,
                            cost_change=cost_change,

                            previous_wholesale_price=old_wholesale,
                            new_wholesale_price=new_wholesale,
                            wholesale_change=wholesale_change,

                            previous_retail_price=old_retail,
                            new_retail_price=new_retail,
                            retail_change=retail_change,
                        )

        except Exception as e:
            raise ValidationError(f"Excel row {excel_row}: {e}")

    return {
        "created": created,
        "updated": updated,
    }
